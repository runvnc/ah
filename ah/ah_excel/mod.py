from ..commands import command
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
import json
import io
from .analyze_excel import analyze_structure

# Module-level cache for the workbook
_workbook_cache = None

def _get_workbook():
    global _workbook_cache
    if _workbook_cache is None:
        raise ValueError("No workbook is currently cached. Use open_workbook first.")
    return load_workbook(filename=io.BytesIO(_workbook_cache))

def _save_workbook_cache(wb):
    global _workbook_cache
    buffer = io.BytesIO()
    wb.save(buffer)
    _workbook_cache = buffer.getvalue()

@command()
async def open_workbook(filename, context=None):
    """Open an Excel workbook and cache it in memory.
    Example:
    { "open_workbook": { "filename": "example.xlsx" } }
    """
    wb = load_workbook(filename)
    _save_workbook_cache(wb)
    context.data['current_workbook'] = filename
    context.data['current_sheet'] = wb.active.title
    return f"Opened and cached workbook {filename}. Sheets: {', '.join(wb.sheetnames)}"

@command()
async def select_sheet(sheet_name, context=None):
    """Select a specific sheet and analyze its structure.
    Example:
    { "select_sheet": { "sheet_name": "Sheet1" } }
    """
    try:
        wb = _get_workbook()
        if sheet_name not in wb.sheetnames:
            return f"Sheet {sheet_name} not found in the workbook."
        
        context.data['current_sheet'] = sheet_name
        structure = analyze_structure(wb[sheet_name])
        
        summary = {
            "headers_count": len(structure["headers"]),
            "data_ranges_count": len(structure["data_ranges"]),
            "empty_columns_count": len(structure["empty_columns"]),
            "empty_rows_count": len(structure["empty_rows"]),
            "merged_cells_count": len(structure["merged_cells"]),
            "info_boxes_count": len(structure["info_boxes"]),
            "text_cells_count": len(structure["text_cells"]),
            "numeric_cells_count": len(structure["numeric_cells"]),
        }
        
        context.data['sheet_structure'] = structure

        print(json.dumps(summary))

        return f"Selected sheet {sheet_name}. Structure summary:\n{json.dumps(summary, indent=2)}\n\nFull structure available in context.data['sheet_structure']."
    except Exception as e:
        print(e)
        return f"Error: {str(e)}"

@command()
async def read_cell(cell_reference, context=None):
    """Read the value of a specific cell.
    Example:
    { "read_cell": { "cell_reference": "B5" } }
    """
    wb = _get_workbook()
    ws = wb[context.data['current_sheet']]
    return str(ws[cell_reference].value)

@command()
async def write_cell(cell_reference, value, context=None):
    """Write a value to a specific cell.
    Example:
    { "write_cell": { "cell_reference": "C7", "value": 1500 } }
    """
    wb = _get_workbook()
    ws = wb[context.data['current_sheet']]
    ws[cell_reference] = value
    _save_workbook_cache(wb)
    return f"Wrote {value} to cell {cell_reference} and updated the cache."

@command()
async def save_workbook(filename=None, context=None):
    """Save the workbook to disk, optionally with a new filename.
    Example:
    { "save_workbook": { "filename": "updated_file.xlsx" } }
    """
    wb = _get_workbook()
    if filename is None:
        filename = context.data['current_workbook']
    wb.save(filename)
    context.data['current_workbook'] = filename
    return f"Saved workbook as {filename}"
